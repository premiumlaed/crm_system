import sys
import pandas as pd
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                            QHBoxLayout, QTabWidget, QPushButton, QLabel, 
                            QFileDialog, QTableWidget, QTableWidgetItem, 
                            QLineEdit, QFormLayout, QComboBox, QMessageBox,
                            QScrollArea, QStyleFactory, QFrame, QTextEdit,
                            QHeaderView, QSizePolicy)
from PyQt5.QtCore import Qt, QTranslator, QLocale
from PyQt5.QtGui import QFont, QIcon, QPalette, QColor

class CRMDashboard(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_language = 'en'
        self.form_fields = {}  # Initialize form_fields dictionary
        self.customer_fields = {
            'basic_info': {
                'name': {'en': 'Full Name', 'ar': 'الاسم الكامل'},
                'email': {'en': 'Email', 'ar': 'البريد الإلكتروني'},
                'phone': {'en': 'Phone', 'ar': 'الهاتف'},
                'alternative_phone': {'en': 'Alternative Phone', 'ar': 'هاتف بديل'},
                'category': {'en': 'Category', 'ar': 'الفئة'},
                'customer_id': {'en': 'Customer ID', 'ar': 'رقم العميل'},
                'registration_date': {'en': 'Registration Date', 'ar': 'تاريخ التسجيل'}
            },
            'social_media': {
                'facebook': {'en': 'Facebook', 'ar': 'فيسبوك'},
                'instagram': {'en': 'Instagram', 'ar': 'انستغرام'},
                'twitter': {'en': 'Twitter', 'ar': 'تويتر'},
                'linkedin': {'en': 'LinkedIn', 'ar': 'لينكد إن'},
                'preferred_contact': {'en': 'Preferred Contact Method', 'ar': 'طريقة التواصل المفضلة'}
            },
            'additional_info': {
                'company': {'en': 'Company', 'ar': 'الشركة'},
                'position': {'en': 'Position', 'ar': 'المنصب'},
                'address': {'en': 'Address', 'ar': 'العنوان'},
                'city': {'en': 'City', 'ar': 'المدينة'},
                'country': {'en': 'Country', 'ar': 'البلد'},
                'notes': {'en': 'Notes', 'ar': 'ملاحظات'}
            }
        }
        self.init_data_storage()
        self.init_translations()
        self.init_ui()
        self.load_saved_data()
    def init_data_storage(self):
        self.data_file = 'crm_data.json'
        self.customers_data = pd.DataFrame()
        self.products_data = pd.DataFrame()

    def init_translations(self):
        self.translations = {
            'en': {
                'window_title': 'CRM Dashboard',
                'customers': 'Customers',
                'products': 'Products',
                'analytics': 'Analytics',
                'add_customer': 'Add Customer',
                'import_customers': 'Import Customers',
                'export_excel': 'Export to Excel',
                'name': 'Name',
                'email': 'Email',
                'phone': 'Phone',
                'category': 'Category',
                'success': 'Success',
                'error': 'Error',
                'warning': 'Warning',
                'download_template': 'Download Template'
            },
            'ar': {
                'window_title': 'لوحة إدارة العملاء',
                'customers': 'العملاء',
                'products': 'المنتجات',
                'analytics': 'التحليلات',
                'add_customer': 'إضافة عميل',
                'import_customers': 'استيراد العملاء',
                'export_excel': 'تصدير إلى Excel',
                'name': 'الاسم',
                'email': 'البريد الإلكتروني',
                'phone': 'الهاتف',
                'category': 'الفئة',
                'success': 'نجاح',
                'error': 'خطأ',
                'warning': 'تحذير',
                'download_template': 'تحميل النموذج'
            }
        }

    def init_ui(self):
        self.setWindowTitle(self.tr_text('window_title'))
        self.setGeometry(100, 100, 1200, 800)
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f0f0;
            }
            QTabWidget::pane {
                border: 2px solid #3498db;
                border-radius: 6px;
                background-color: #ffffff;
            }
            QTabBar::tab {
                background-color: #3498db;
                color: white;
                padding: 10px 20px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: #2980b9;
            }
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 4px;
                min-width: 100px;
                margin: 2px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #2573a7;
            }
            QTableWidget {
                border: 1px solid #bdc3c7;
                gridline-color: #ecf0f1;
                border-radius: 4px;
                alternate-background-color: #f9f9f9;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 5px;
                border: none;
            }
            QLineEdit {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
            }
            QLineEdit:focus {
                border: 2px solid #3498db;
            }
            QComboBox {
                padding: 8px;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
            }
            QComboBox:drop-down {
                border: none;
            }
            QComboBox:down-arrow {
                image: url(down_arrow.png);
                width: 12px;
                height: 12px;
            }
            QTextEdit {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                background-color: white;
                padding: 5px;
            }
            QFrame {
                background-color: white;
                border-radius: 6px;
                padding: 10px;
            }
            QLabel {
                color: #2c3e50;
            }
        """)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)

        # Language switcher
        language_layout = QHBoxLayout()
        language_btn = QPushButton("🌐 English/عربي")
        language_btn.clicked.connect(self.toggle_language)
        language_layout.addWidget(language_btn, alignment=Qt.AlignRight)
        main_layout.addLayout(language_layout)

        # Create tab widget
        self.tab_widget = QTabWidget()
        self.customers_tab = self.create_customers_tab()
        self.products_tab = self.create_products_tab()
        self.analytics_tab = self.create_analytics_tab()

        self.tab_widget.addTab(self.customers_tab, self.tr_text('customers'))
        self.tab_widget.addTab(self.products_tab, self.tr_text('products'))
        self.tab_widget.addTab(self.analytics_tab, self.tr_text('analytics'))

        main_layout.addWidget(self.tab_widget)
    def create_customers_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()

        # Create button group with spacing and alignment
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)
        
        # Add Customer button
        add_customer_btn = QPushButton(self.tr_text('add_customer'))
        add_customer_btn.clicked.connect(self.show_add_customer_form)
        buttons_layout.addWidget(add_customer_btn)

        # Import Customers button
        import_customers_btn = QPushButton(self.tr_text('import_customers'))
        import_customers_btn.clicked.connect(self.import_customers)
        buttons_layout.addWidget(import_customers_btn)

        # Export Excel button
        export_excel_btn = QPushButton(self.tr_text('export_excel'))
        export_excel_btn.clicked.connect(self.export_customers_excel)
        buttons_layout.addWidget(export_excel_btn)

        # Template button
        template_btn = QPushButton(self.tr_text('download_template'))
        template_btn.clicked.connect(lambda: self.export_template('customers'))
        buttons_layout.addWidget(template_btn)

        # Add stretch to push buttons to the left
        buttons_layout.addStretch()
        layout.addLayout(buttons_layout)

        # Create and setup customers table
        self.customers_table = QTableWidget()
        self.customers_table.setAlternatingRowColors(True)
        self.customers_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.customers_table.setSelectionMode(QTableWidget.SingleSelection)
        self.customers_table.horizontalHeader().setStretchLastSection(True)
        self.customers_table.verticalHeader().setVisible(False)
        layout.addWidget(self.customers_table)

        tab.setLayout(layout)
        return tab

    def create_products_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Create button group
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(10)
        
        # Import Products button
        import_products_btn = QPushButton("Import Products")
        import_products_btn.clicked.connect(self.import_products)
        buttons_layout.addWidget(import_products_btn)
        
        # Export Excel button
        export_excel_btn = QPushButton("Export to Excel")
        export_excel_btn.clicked.connect(self.export_products_excel)
        buttons_layout.addWidget(export_excel_btn)
        
        # Template button
        template_btn = QPushButton(self.tr_text('download_template'))
        template_btn.clicked.connect(lambda: self.export_template('products'))
        buttons_layout.addWidget(template_btn)
        
        # Add stretch to push buttons to the left
        buttons_layout.addStretch()
        layout.addLayout(buttons_layout)
        
        # Create and setup products table
        self.products_table = QTableWidget()
        self.products_table.setAlternatingRowColors(True)
        self.products_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.products_table.setSelectionMode(QTableWidget.SingleSelection)
        self.products_table.horizontalHeader().setStretchLastSection(True)
        self.products_table.verticalHeader().setVisible(False)
        layout.addWidget(self.products_table)
        
        tab.setLayout(layout)
        return tab

    def create_analytics_tab(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        # Create analytics dashboard
        analytics_frame = QFrame()
        analytics_frame.setFrameStyle(QFrame.StyledPanel)
        analytics_layout = QVBoxLayout(analytics_frame)
        
        # Add analytics title
        title_label = QLabel("Analytics Dashboard")
        title_label.setStyleSheet("""
            font-size: 18px;
            font-weight: bold;
            color: #2c3e50;
            margin: 10px;
        """)
        title_label.setAlignment(Qt.AlignCenter)
        analytics_layout.addWidget(title_label)
        
        # Add placeholder for future analytics features
        placeholder_label = QLabel("Analytics features coming soon...")
        placeholder_label.setAlignment(Qt.AlignCenter)
        placeholder_label.setStyleSheet("color: #7f8c8d; font-size: 14px;")
        analytics_layout.addWidget(placeholder_label)
        
        layout.addWidget(analytics_frame)
        tab.setLayout(layout)
        return tab

    def show_add_customer_form(self):
        self.customer_form = QWidget()
        self.customer_form.setWindowTitle(self.tr_text('add_customer'))
        self.customer_form.setMinimumWidth(600)
        self.customer_form.setWindowModality(Qt.ApplicationModal)  # Make form modal
        
        main_layout = QVBoxLayout()
        scroll = QScrollArea()
        scroll_widget = QWidget()
        form_layout = QFormLayout()

        # Clear previous form fields
        self.form_fields = {}

        # Create sections
        sections = {
            'basic_info': self.create_section('Basic Information', self.customer_fields['basic_info']),
            'social_media': self.create_section('Social Media', self.customer_fields['social_media']),
            'additional_info': self.create_section('Additional Information', self.customer_fields['additional_info'])
        }

        for section in sections.values():
            form_layout.addRow(section)

        # Add submit button
        submit_btn = QPushButton(self.tr_text('add_customer'))
        submit_btn.clicked.connect(self.add_customer)
        form_layout.addRow(submit_btn)

        scroll_widget.setLayout(form_layout)
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        main_layout.addWidget(scroll)
        
        self.customer_form.setLayout(main_layout)
        self.customer_form.show()
    def create_section(self, title, fields):
        section = QFrame()
        section.setFrameStyle(QFrame.Box | QFrame.Raised)
        layout = QFormLayout()
        
        title_label = QLabel(title)
        title_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #2c3e50;")
        layout.addRow(title_label)

        for field_key, field_labels in fields.items():
            if field_key == 'category':
                self.form_fields[field_key] = QComboBox()
                self.form_fields[field_key].addItems([
                    "Real Estate", "Tourism", "E-store", "Retail", "Other"
                ])
            elif field_key == 'preferred_contact':
                self.form_fields[field_key] = QComboBox()
                self.form_fields[field_key].addItems([
                    "Email", "Phone", "Facebook", "Instagram", "Twitter", "LinkedIn"
                ])
            elif field_key == 'notes':
                self.form_fields[field_key] = QTextEdit()
                self.form_fields[field_key].setMaximumHeight(100)
            else:
                self.form_fields[field_key] = QLineEdit()

            layout.addRow(field_labels['en'], self.form_fields[field_key])

        section.setLayout(layout)
        return section

    def create_sample_template(self, template_type='customers'):
        if template_type == 'customers':
            sample_data = {
                'name': ['John Doe', 'Jane Smith'],
                'email': ['john@example.com', 'jane@example.com'],
                'phone': ['+1234567890', '+0987654321'],
                'alternative_phone': ['+1122334455', '+5544332211'],
                'category': ['Real Estate', 'Tourism'],
                'facebook': ['fb.com/john', 'fb.com/jane'],
                'instagram': ['@john_doe', '@jane_smith'],
                'twitter': ['@johnd', '@janes'],
                'linkedin': ['linkedin.com/john', 'linkedin.com/jane'],
                'preferred_contact': ['Email', 'Phone'],
                'company': ['ABC Corp', 'XYZ Ltd'],
                'position': ['Manager', 'Director'],
                'address': ['123 Main St', '456 Oak Ave'],
                'city': ['New York', 'Los Angeles'],
                'country': ['USA', 'USA'],
                'notes': ['VIP Customer', 'Regular Customer']
            }
        else:  # products template
            sample_data = {
                'product_id': ['PRD001', 'PRD002'],
                'name': ['Product 1', 'Product 2'],
                'category': ['Category A', 'Category B'],
                'price': [99.99, 149.99],
                'stock': [100, 50],
                'description': ['Product 1 description', 'Product 2 description'],
                'supplier': ['Supplier A', 'Supplier B'],
                'status': ['Active', 'Active']
            }
        
        df = pd.DataFrame(sample_data)
        return df

    def export_template(self, template_type='customers'):
        try:
            template_df = self.create_sample_template(template_type)
            
            file_name, _ = QFileDialog.getSaveFileName(
                self, f"Save {template_type.capitalize()} Template", 
                f"{template_type}_template.xlsx", 
                "Excel Files (*.xlsx)")
            
            if file_name:
                wb = Workbook()
                ws = wb.active
                ws.title = f"{template_type.capitalize()} Template"

                # Style the header
                header_fill = PatternFill(start_color="1F497D", 
                                        end_color="1F497D", 
                                        fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                header_alignment = Alignment(horizontal="center", 
                                          vertical="center")
                border = Border(left=Side(style='thin'), 
                              right=Side(style='thin'),
                              top=Side(style='thin'), 
                              bottom=Side(style='thin'))

                # Write headers
                for col, header in enumerate(template_df.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border

                # Write sample data
                for row, data in enumerate(template_df.values, 2):
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left")

                # Add instructions sheet
                ws_instructions = wb.create_sheet("Instructions")
                instructions = [
                    "Instructions for using this template:",
                    "",
                    f"1. This is a sample {template_type} template with example data",
                    "2. Replace the example data with your actual data",
                    "3. Keep the column headers exactly as they are",
                    "4. Save the file and use it to import into the CRM system",
                    "",
                    "Required fields:",
                ]
                
                if template_type == 'customers':
                    required_fields = ["name", "email", "phone"]
                else:
                    required_fields = ["product_id", "name", "category"]
                
                for idx, instruction in enumerate(instructions, 1):
                    ws_instructions.cell(row=idx, column=1, value=instruction)
                
                for idx, field in enumerate(required_fields, len(instructions) + 2):
                    ws_instructions.cell(row=idx, column=1, value=f"- {field}")

                # Adjust column widths
                for ws in [ws, ws_instructions]:
                    for column in ws.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column[0].column_letter].width = adjusted_width

                wb.save(file_name)
                QMessageBox.information(self, "Success", 
                                      f"{template_type.capitalize()} template created successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error creating template: {str(e)}")
    def add_customer(self):
        try:
            # Validate required fields
            required_fields = ['name', 'email', 'phone']
            for field in required_fields:
                if not self.form_fields[field].text().strip():
                    QMessageBox.warning(self, self.tr_text('warning'),
                                      f"{field.capitalize()} is required!")
                    return

            # Validate email format
            email = self.form_fields['email'].text()
            if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
                QMessageBox.warning(self, self.tr_text('warning'),
                                  "Invalid email format!")
                return

            # Create new customer dictionary
            new_customer = {}
            for section in self.customer_fields.values():
                for field_key in section:
                    if isinstance(self.form_fields[field_key], QComboBox):
                        new_customer[field_key] = self.form_fields[field_key].currentText()
                    elif isinstance(self.form_fields[field_key], QTextEdit):
                        new_customer[field_key] = self.form_fields[field_key].toPlainText()
                    else:
                        new_customer[field_key] = self.form_fields[field_key].text()

            # Add timestamp and customer ID
            new_customer['registration_date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            new_customer['customer_id'] = f"CUS{len(self.customers_data) + 1:06d}"

            # Add to DataFrame
            self.customers_data = pd.concat([self.customers_data, 
                                           pd.DataFrame([new_customer])], 
                                          ignore_index=True)
            
            # Update table and save data
            self.update_customers_table()
            self.save_data()
            
            # Close form and show success message
            self.customer_form.close()
            QMessageBox.information(self, self.tr_text('success'), 
                                  "Customer added successfully!")
                                  
        except Exception as e:
            QMessageBox.critical(self, self.tr_text('error'),
                               f"Error adding customer: {str(e)}")

    def import_customers(self):
        try:
            file_name, _ = QFileDialog.getOpenFileName(
                self, "Import Customers", "", 
                "Excel Files (*.xlsx *.xls);;CSV Files (*.csv)")
            if file_name:
                if file_name.endswith(('.xlsx', '.xls')):
                    imported_data = pd.read_excel(file_name)
                else:
                    imported_data = pd.read_csv(file_name)
                
                # Validate required columns
                required_columns = ['name', 'email', 'phone']
                missing_columns = [col for col in required_columns if col not in imported_data.columns]
                if missing_columns:
                    QMessageBox.warning(self, self.tr_text('warning'),
                                      f"Missing required columns: {', '.join(missing_columns)}")
                    return

                # Add customer IDs and registration dates if not present
                if 'customer_id' not in imported_data.columns:
                    start_id = len(self.customers_data) + 1
                    imported_data['customer_id'] = [f"CUS{i:06d}" for i in range(start_id, start_id + len(imported_data))]
                
                if 'registration_date' not in imported_data.columns:
                    imported_data['registration_date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Concatenate with existing data
                self.customers_data = pd.concat([self.customers_data, imported_data], 
                                              ignore_index=True)
                
                self.update_customers_table()
                self.save_data()
                QMessageBox.information(self, self.tr_text('success'),
                                      "Customers imported successfully!")
        except Exception as e:
            QMessageBox.critical(self, self.tr_text('error'),
                               f"Error importing file: {str(e)}")

    def import_products(self):
        try:
            file_name, _ = QFileDialog.getOpenFileName(
                self, "Import Products", "", 
                "Excel Files (*.xlsx *.xls);;CSV Files (*.csv)")
            if file_name:
                if file_name.endswith(('.xlsx', '.xls')):
                    imported_data = pd.read_excel(file_name)
                else:
                    imported_data = pd.read_csv(file_name)
                
                # Validate required columns
                required_columns = ['product_id', 'name', 'category']
                missing_columns = [col for col in required_columns if col not in imported_data.columns]
                if missing_columns:
                    QMessageBox.warning(self, self.tr_text('warning'),
                                      f"Missing required columns: {', '.join(missing_columns)}")
                    return

                # Concatenate with existing data
                self.products_data = pd.concat([self.products_data, imported_data], 
                                             ignore_index=True)
                
                self.update_products_table()
                self.save_data()
                QMessageBox.information(self, self.tr_text('success'),
                                      "Products imported successfully!")
        except Exception as e:
            QMessageBox.critical(self, self.tr_text('error'),
                               f"Error importing file: {str(e)}")

    def export_customers_excel(self):
        if self.customers_data.empty:
            QMessageBox.warning(self, self.tr_text('warning'), 
                              "No customer data to export!")
            return

        try:
            file_name, _ = QFileDialog.getSaveFileName(
                self, "Export Customers", "", 
                "Excel Files (*.xlsx)")
            
            if file_name:
                wb = Workbook()
                ws = wb.active
                ws.title = "Customer Data"         
         # Style definitions
                header_fill = PatternFill(start_color="1F497D", 
                                        end_color="1F497D", 
                                        fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                header_alignment = Alignment(horizontal="center", 
                                          vertical="center")
                border = Border(left=Side(style='thin'), 
                              right=Side(style='thin'),
                              top=Side(style='thin'), 
                              bottom=Side(style='thin'))

                # Write headers with styling
                for col, header in enumerate(self.customers_data.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border

                # Write data
                for row, data in enumerate(self.customers_data.values, 2):
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

                wb.save(file_name)
                QMessageBox.information(self, self.tr_text('success'),
                                      "Customers exported successfully!")
        except Exception as e:
            QMessageBox.critical(self, self.tr_text('error'),
                               f"Error exporting file: {str(e)}")

    def export_products_excel(self):
        if self.products_data.empty:
            QMessageBox.warning(self, self.tr_text('warning'), 
                              "No product data to export!")
            return

        try:
            file_name, _ = QFileDialog.getSaveFileName(
                self, "Export Products", "", 
                "Excel Files (*.xlsx)")
            
            if file_name:
                wb = Workbook()
                ws = wb.active
                ws.title = "Product Data"

                # Style definitions
                header_fill = PatternFill(start_color="1F497D", 
                                        end_color="1F497D", 
                                        fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                header_alignment = Alignment(horizontal="center", 
                                          vertical="center")
                border = Border(left=Side(style='thin'), 
                              right=Side(style='thin'),
                              top=Side(style='thin'), 
                              bottom=Side(style='thin'))

                # Write headers with styling
                for col, header in enumerate(self.products_data.columns, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border

                # Write data
                for row, data in enumerate(self.products_data.values, 2):
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

                wb.save(file_name)
                QMessageBox.information(self, self.tr_text('success'),
                                      "Products exported successfully!")
        except Exception as e:
            QMessageBox.critical(self, self.tr_text('error'),
                               f"Error exporting file: {str(e)}")

    def update_customers_table(self):
        try:
            self.customers_table.setRowCount(len(self.customers_data))
            self.customers_table.setColumnCount(len(self.customers_data.columns))
            self.customers_table.setHorizontalHeaderLabels(self.customers_data.columns)

            header = self.customers_table.horizontalHeader()
            for i in range(len(self.customers_data.columns)):
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)

            for i in range(len(self.customers_data)):
                for j in range(len(self.customers_data.columns)):
                    item = QTableWidgetItem(str(self.customers_data.iloc[i, j]))
                    self.customers_table.setItem(i, j, item)

            self.customers_table.setSortingEnabled(True)
        except Exception as e:
            QMessageBox.critical(self, self.tr_text('error'),
                               f"Error updating table: {str(e)}")
    def update_products_table(self):
        try:
            self.products_table.setRowCount(len(self.products_data))
            self.products_table.setColumnCount(len(self.products_data.columns))
            self.products_table.setHorizontalHeaderLabels(self.products_data.columns)

            header = self.products_table.horizontalHeader()
            for i in range(len(self.products_data.columns)):
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)

            for i in range(len(self.products_data)):
                for j in range(len(self.products_data.columns)):
                    item = QTableWidgetItem(str(self.products_data.iloc[i, j]))
                    self.products_table.setItem(i, j, item)

            self.products_table.setSortingEnabled(True)
        except Exception as e:
            QMessageBox.critical(self, self.tr_text('error'),
                               f"Error updating table: {str(e)}")

    def save_data(self):
        try:
            data = {
                'customers': self.customers_data.to_dict('records'),
                'products': self.products_data.to_dict('records')
            }
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            QMessageBox.critical(self, self.tr_text('error'),
                               f"Error saving data: {str(e)}")

    def load_saved_data(self):
        try:
            if os.path.exists(self.data_file):
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.customers_data = pd.DataFrame(data.get('customers', []))
                    self.products_data = pd.DataFrame(data.get('products', []))
                    self.update_customers_table()
                    self.update_products_table()
        except Exception as e:
            QMessageBox.warning(self, self.tr_text('warning'),
                              f"Error loading data: {str(e)}")

    def tr_text(self, key):
        return self.translations[self.current_language].get(key, key)

    def toggle_language(self):
        self.current_language = 'ar' if self.current_language == 'en' else 'en'
        self.retranslate_ui()

    def retranslate_ui(self):
        self.setWindowTitle(self.tr_text('window_title'))
        self.tab_widget.setTabText(0, self.tr_text('customers'))
        self.tab_widget.setTabText(1, self.tr_text('products'))
        self.tab_widget.setTabText(2, self.tr_text('analytics'))

    def closeEvent(self, event):
        try:
            self.save_data()
            event.accept()
        except Exception as e:
            QMessageBox.critical(self, self.tr_text('error'),
                               f"Error saving data before exit: {str(e)}")
            event.accept()

if __name__ == '__main__':
    try:
        app = QApplication(sys.argv)
        app.setStyle('Fusion')
        font = QFont('Arial', 10)
        app.setFont(font)
        window = CRMDashboard()
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        print(f"Critical error: {str(e)}")
